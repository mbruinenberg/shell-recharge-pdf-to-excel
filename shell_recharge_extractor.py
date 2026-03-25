#!/usr/bin/env python3
"""
Shell Recharge Invoice Extractor
Extracts charging session data from Shell Recharge PDF receipts
and outputs a summary Excel file.

Usage:
    python shell_recharge_extractor.py <pdf_folder> [output_file.xlsx]

Example:
    python shell_recharge_extractor.py ./test/invoices
    python shell_recharge_extractor.py ./test/invoices output.xlsx
"""

import sys
import os
import re
import glob
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


def extract_receipt_data(pdf_path):
    """Extract structured data from a Shell Recharge PDF receipt."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
    except Exception as e:
        print(f"  ERROR reading {pdf_path}: {e}")
        return None

    data = {"source_file": os.path.basename(pdf_path)}

    # Receipt number
    m = re.search(r'Receipt\s*#\s*(\S+)', text)
    data["receipt_number"] = m.group(1) if m else ""

    # Issue date
    m = re.search(r'Issue\s*Date:\s*(\d{2}/\d{2}/\d{4})', text)
    data["issue_date"] = m.group(1) if m else ""

    # Charging session start date/time
    m = re.search(r'Start\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})', text)
    if m:
        data["session_start_date"] = m.group(1)
        data["session_start_time"] = m.group(2)
        try:
            data["session_start_dt"] = datetime.strptime(
                f"{m.group(1)} {m.group(2)}", "%d/%m/%Y %H:%M"
            )
        except ValueError:
            data["session_start_dt"] = None
    else:
        data["session_start_date"] = ""
        data["session_start_time"] = ""
        data["session_start_dt"] = None

    # Charging session end date/time
    m = re.search(r'End\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})', text)
    if m:
        data["session_end_date"] = m.group(1)
        data["session_end_time"] = m.group(2)
    else:
        data["session_end_date"] = ""
        data["session_end_time"] = ""

    # Duration
    m = re.search(r'Duration:\s*([\d:]+)', text)
    data["duration"] = m.group(1) if m else ""

    # Energy (kWh)
    m = re.search(r'Energy:\s*([\d.,]+)\s*kWh', text)
    data["energy_kwh"] = float(m.group(1).replace(",", ".")) if m else 0.0

    # Price per kWh
    m = re.search(r'Price\s*per\s*kWh:\s*([\d.,]+)\s*EUR', text)
    data["price_per_kwh"] = float(m.group(1).replace(",", ".")) if m else 0.0

    # Transaction fee
    m = re.search(r'Transaction\s*fee:\s*([\d.,]+)\s*EUR', text)
    data["transaction_fee"] = float(m.group(1).replace(",", ".")) if m else 0.0

    # Station info (location)
    m = re.search(r'Charging\s+Session\s+(.+?)\s+\d+[\.,]\d+\s+EUR', text)
    data["station"] = m.group(1).strip() if m else ""

    # Station city - line after the station address
    m = re.search(r'Start\s+\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}\s+(.+)', text)
    data["station_city"] = m.group(1).strip() if m else ""

    # Amount before VAT
    m = re.search(r'Amount\s+before\s+VAT\s+([\d.,]+)\s*EUR', text)
    data["amount_before_vat"] = float(m.group(1).replace(",", ".")) if m else 0.0

    # VAT percentage and amount
    m = re.search(r'VAT\s+Total\s+\(([\d.,]+)%\)\s+([\d.,]+)\s*EUR', text)
    if m:
        data["vat_percentage"] = float(m.group(1).replace(",", "."))
        data["vat_amount"] = float(m.group(2).replace(",", "."))
    else:
        data["vat_percentage"] = 0.0
        data["vat_amount"] = 0.0

    # Amount including VAT
    m = re.search(r'Amount\s+incl\.\s*VAT\s+([\d.,]+)\s*EUR', text)
    data["amount_incl_vat"] = float(m.group(1).replace(",", ".")) if m else 0.0

    # Payment method
    m = re.search(r'Payment\s+Method\s+(.+)', text)
    data["payment_method"] = m.group(1).strip() if m else ""

    return data


def create_excel(records, output_path):
    """Create a formatted Excel workbook from extracted records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shell Recharge Invoices"

    # Column definitions: (header, data_key, width, number_format)
    columns = [
        ("Receipt #", "receipt_number", 20, None),
        ("Charge Date", "session_start_dt", 14, "DD-MM-YYYY"),
        ("Start Time", "session_start_time", 12, None),
        ("End Time", "session_end_time", 12, None),
        ("Duration", "duration", 12, None),
        ("Station", "station", 22, None),
        ("City", "station_city", 16, None),
        ("Energy (kWh)", "energy_kwh", 14, '#,##0.00'),
        ("Price/kWh (EUR)", "price_per_kwh", 16, '€ #,##0.00'),
        ("Transaction Fee", "transaction_fee", 16, '€ #,##0.00'),
        ("Amount excl. VAT", "amount_before_vat", 18, '€ #,##0.00'),
        ("VAT %", "vat_percentage", 10, '0.00"%"'),
        ("VAT Amount", "vat_amount", 14, '€ #,##0.00'),
        ("Amount incl. VAT", "amount_incl_vat", 18, '€ #,##0.00'),
        ("Payment Method", "payment_method", 18, None),
        ("Source File", "source_file", 30, None),
    ]

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="404040")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    eur_format = '€ #,##0.00'

    # Write headers
    for col_idx, (header, _, width, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, record in enumerate(records, 2):
        for col_idx, (_, key, _, fmt) in enumerate(columns, 1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if fmt:
                cell.number_format = fmt

    # Summary rows
    if records:
        last_data_row = len(records) + 1
        summary_row = last_data_row + 2

        ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(
            name="Arial", bold=True, size=10
        )

        # Sum formulas for financial columns
        sum_cols = {
            "amount_before_vat": 11,
            "vat_amount": 13,
            "amount_incl_vat": 14,
        }
        for key, col_idx in sum_cols.items():
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=summary_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{last_data_row})"
            )
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.number_format = eur_format

        # Count
        ws.cell(row=summary_row, column=2, value=f"=COUNTA(B2:B{last_data_row})")
        ws.cell(row=summary_row, column=2).font = Font(name="Arial", bold=True, size=10)

        # Total energy
        col_letter = get_column_letter(8)
        cell = ws.cell(
            row=summary_row, column=8,
            value=f"=SUM({col_letter}2:{col_letter}{last_data_row})"
        )
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.number_format = '#,##0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python shell_recharge_extractor.py <pdf_folder> [output.xlsx]")
        print('Example: python shell_recharge_extractor.py "./Shell Recharge"')
        sys.exit(1)

    pdf_folder = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        pdf_folder, f"Shell_Recharge_Summary_{datetime.now().strftime('%Y-%m')}.xlsx"
    )

    if not os.path.isdir(pdf_folder):
        print(f"Error: Folder not found: {pdf_folder}")
        sys.exit(1)

    pdf_files = sorted(glob.glob(os.path.join(pdf_folder, "*.pdf")))
    if not pdf_files:
        print(f"No PDF files found in: {pdf_folder}")
        sys.exit(1)

    print(f"Found {len(pdf_files)} PDF file(s) in {pdf_folder}")
    records = []
    for pdf_path in pdf_files:
        print(f"  Processing: {os.path.basename(pdf_path)}")
        data = extract_receipt_data(pdf_path)
        if data:
            records.append(data)

    if not records:
        print("No data extracted.")
        sys.exit(1)

    # Sort by session start date
    records.sort(key=lambda r: r.get("session_start_dt") or datetime.min)

    create_excel(records, output_file)
    print(f"\nDone! {len(records)} receipt(s) → {output_file}")


if __name__ == "__main__":
    main()
